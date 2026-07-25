#!/usr/bin/env python3
"""Build Excel review workbook from categorize-dump-for-approval JSON dump."""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path.home() / "OneDrive" / "Desktop" / "AI accounting" / "2026_Uncategorized_Categorization_Review.xlsx"


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else "scripts/reconciliation/output/_cat_approve_2026.json")
    data = json.loads(src.read_text(encoding="utf-8"))
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "2026 Uncategorized → Categorization for review & approval"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Scanned dump lines", data.get("scanned")),
        ("Draft categorizations created", data.get("draftsCreated")),
        ("Still need a rule / Jerry pick", data.get("needsReviewCount")),
        ("Source PDFs attached", data.get("documentsAttached")),
        ("Approve drafts in app", "https://ljc-accounting-app.onrender.com/draft-journals?all=1"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        if isinstance(v, str) and v.startswith("http"):
            ws[f"B{i}"].hyperlink = v
            ws[f"B{i}"].font = Font(color="0563C1", underline="single")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 70

    def fill_sheet(name, items, headers, row_fn):
        sh = wb.create_sheet(name)
        head_fill = PatternFill("solid", fgColor="1F4E79")
        head_font = Font(color="FFFFFF", bold=True)
        for c, h in enumerate(headers, 1):
            cell = sh.cell(1, c, h)
            cell.fill = head_fill
            cell.font = head_font
        for r, item in enumerate(items or [], 2):
            vals = row_fn(item)
            for c, v in enumerate(vals, 1):
                cell = sh.cell(r, c, v)
                if isinstance(v, str) and (v.startswith("http") or v.startswith("file:")):
                    cell.hyperlink = v
                    cell.font = Font(color="0563C1", underline="single")
        for c in range(1, len(headers) + 1):
            sh.column_dimensions[get_column_letter(c)].width = 22
        sh.column_dimensions["E"].width = 55
        sh.column_dimensions["H"].width = 45

    fill_sheet(
        "Drafts to approve",
        data.get("proposed") or [],
        ["Date", "Amount", "From", "To", "Description", "Source JE", "Draft JE", "Open draft", "Stmt date", "Doc"],
        lambda x: [
            x.get("postingDate"),
            x.get("amount"),
            x.get("fromAccount"),
            x.get("toAccount"),
            x.get("description"),
            x.get("jeNumber"),
            x.get("draftJeNumber"),
            x.get("approveUrl") or x.get("appUrl"),
            x.get("statementDate"),
            (x.get("document") or {}).get("fileName") or ("yes" if (x.get("document") or {}).get("alreadyAttached") else ""),
        ],
    )

    fill_sheet(
        "Still need your pick",
        data.get("needsReview") or [],
        ["Date", "Amount", "From", "Description", "Source JE", "Open JE", "Stmt date", "Why"],
        lambda x: [
            x.get("postingDate"),
            x.get("amount"),
            x.get("fromAccount"),
            x.get("description"),
            x.get("jeNumber"),
            x.get("appUrl"),
            x.get("statementDate"),
            x.get("reason"),
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print("WROTE", OUT)


if __name__ == "__main__":
    main()
